from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image as PilImage
from PIL import ImageDraw


ROOT = Path(__file__).resolve().parents[1]
FIXTURE_DIR = ROOT / "tests" / "fixtures" / "clean"
ASSET_DIR = ROOT / "tests" / "fixtures" / "_assets"


def save_basic() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["Date", "Product", "Amount"])
    ws.append(["2026-04-22", "A", 128])
    ws.append(["2026-04-23", "B", 256])
    ws["E2"] = "=SUM(C2:C3)"
    wb.save(FIXTURE_DIR / "normal_basic.xlsx")


def save_validation() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation"
    ws["A1"] = "Status"
    dv = DataValidation(type="list", formula1='"Pending,Done,Cancelled"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("A2:A20")
    ws["A2"] = "Pending"
    wb.save(FIXTURE_DIR / "normal_validation.xlsx")


def save_chart() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Chart"
    rows = [
        ["Month", "Orders"],
        ["Jan", 10],
        ["Feb", 18],
        ["Mar", 26],
    ]
    for row in rows:
        ws.append(row)
    chart = BarChart()
    chart.title = "Orders"
    data = Reference(ws, min_col=2, min_row=1, max_row=4)
    cats = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")
    wb.save(FIXTURE_DIR / "normal_chart.xlsx")


def save_image() -> None:
    ASSET_DIR.mkdir(parents=True, exist_ok=True)
    image_path = ASSET_DIR / "sample.png"
    canvas = PilImage.new("RGB", (160, 80), "#E7F5F2")
    draw = ImageDraw.Draw(canvas)
    draw.rectangle((8, 8, 152, 72), outline="#0B7D69", width=3)
    draw.text((24, 30), "Excel-Dr", fill="#102027")
    canvas.save(image_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Image"
    ws["A1"] = "Normal visible image"
    ws.add_image(Image(str(image_path)), "C3")
    wb.save(FIXTURE_DIR / "normal_image.xlsx")


def main() -> int:
    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
    save_basic()
    save_validation()
    save_chart()
    save_image()
    for path in sorted(FIXTURE_DIR.glob("*.xlsx")):
        print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
